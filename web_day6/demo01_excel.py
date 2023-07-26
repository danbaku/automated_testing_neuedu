"""
作者：张浩
时间：2023/7/25  10:31
邮箱：88302599@qq.com
"""
"""
pip install openpyxl
"""


import openpyxl

#---------------read_excel------------------
#
#1、加载excel文件
workbook=openpyxl.load_workbook(r"E:\class_2017\web_day6\test.xlsx")
# print(workbook.sheetnames)

#2、获取指定sheet的对象
sh = workbook['Sheet1']
# print(sh)

#3-1、单元格读取数据
# cell_res = sh.cell(row= 2,column=2)
# res = cell_res.value
# print(res)

#----------------------读取sheet页面中所有的数据----------------------------------

#3-2、读取所有行的数据
# res = sh.rows #所有行的对象

res = list(sh.rows)
# print(res)

for row in  res:
    for cell in row:
        print(cell.value," ",end="")
    print("")


#----------------写入excel文件---------------------------------

# #1、加载excel文件
# workbook=openpyxl.load_workbook(r"E:\class_2017\web_day6\test.xlsx")
# # print(workbook.sheetnames)
#
# #2、获取指定sheet的对象
# sh = workbook['Sheet1']
# # print(sh)
#
#
# #3、根据单元格的定位写入内容
# sh.cell(row=4,column=4,value="excel_write")
#
# #4、保存文件
# workbook.save(r"E:\class_2017\web_day6\test.xlsx")